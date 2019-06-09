import onlineapp
import  click
import openpyxl
import os
import django
import classproject


os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'classproject.settings')
django.setup()


def importcollege(worksheet):
    flag = 0
    for row in worksheet.rows:
        if (flag == 0):
            flag = 1
        else:
            try:
                college = onlineapp.models.College(name=row[0].value, acronym=row[1].value, location=row[2].value,
                                                   contact=row[3].value)
                print(college)
                college.save()
            except:
                print("Failed Tuple:", row)


def importstudent(worksheet):
    flag = 0
    for row in worksheet.rows:
        if (flag == 0):
            flag = 1
        else:
            try:
                college_data = onlineapp.models.College.objects.get(acronym=row[1].value)
                print(college_data)
                db_folder_name = "ol2016_%s_%s_mock" % (row[1].value, row[-1].value.lower())
                student = onlineapp.models.Student(name=row[0].value, college=college_data, email=row[2].value,
                                                   db_folder=db_folder_name)
                student.save()
            except:
                print("Failed Tuple:", row)

def importmocktest(worksheet):
    skip_flag = True
    for row in worksheet.iter_rows(values_only=True):
        if skip_flag:
            skip_flag = False
            continue
        try:
            student_data =onlineapp.models.Student.objects.get(db_folder=row[0])
            mocktest1 = onlineapp.models.MockTest1(student=student_data, problem1=row[1], problem2=row[2], problem3=row[3],
                                  problem4=row[4], total=row[5])
            mocktest1.save()
        except:
            print("Failed tuple:", row)


def importdropedout(worksheet):
    flag = 0
    for row in worksheet.rows:
        if (flag == 0):
            flag = 1
        else:
            try:
                college_data = onlineapp.models.College.objects.get(acronym=row[1].value)
                print(college_data)
                db_folder_name = "ol2016_%s_%s_mock" % (row[1].value, row[-1].value.lower())
                student = onlineapp.models.Student(name=row[0].value, college=college_data, email=row[2].value,
                                                   db_folder=db_folder_name,dropped_out=True)
                student.save()
            except:
                print("Failed Tuple:", row)


@click.command()
@click.argument('source')
def importdata(source):
    wb = openpyxl.load_workbook(source)
    importcollege(wb["Colleges"])
    print("colleges are imported success fully ")
    importstudent(wb["Current"])
    print("students are imported success fully ")
    importdropedout(wb['Deletions'])
    print("dropped out students are successfully")
    wb=openpyxl.load_workbook('mockresult.xlsx')
    importmocktest(wb.active)
    print("marksdone are imported success fully ")

if __name__=='__main__':
    importdata()