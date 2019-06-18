import iplapp
import  click
import openpyxl
import os
import django
import iplproject


os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'iplproject.settings')
django.setup()


def importmatches(worksheet):
    skip_flag = True
    for row in worksheet.iter_rows(values_only=True):
        if skip_flag:
            skip_flag = False
            continue
        if (row[0] is not None ):
            match = iplapp.models.Matches(
                m_id=row[0],
                season=row[1],
                city=row[2],
                date=row[3],
                team1=row[4],
                team2=row[5],
                toss_winner=row[6],
                toss_decision=row[7],
                result=row[8],
                dl_applied=row[9],
                winner=row[10],
                win_by_runs=row[11],
                win_by_wickets=row[12],
                player_of_match=row[13],
                venue=row[14],
                umpire1=row[15],
                umpire2=row[16],
                umpire3=row[17],
            )
        match.save()
        # try:
        #     if(row[0]!=''):
        #         match = iplapp.models.Matches(
        #         m_id=row[0],
        #         season =row[1],
        #         city = row[2],
        #         date = row[3],
        #         team1 = row[4],
        #         team2 = row[5],
        #         toss_winner = row[6],
        #         toss_decision = row[7],
        #         result = row[8],
        #         dl_applied = row[9],
        #         winner = row[10],
        #         win_by_runs = row[11],
        #         win_by_wickets = row[12],
        #         player_of_match = row[13],
        #         venue = row[14],
        #         umpire1 = row[15],
        #         umpire2 = row[16],
        #         umpire3 =row[17],
        #     )
        #     match.save()
        # except:
        #     print("Failed tuple:", row)


@click.command()
@click.argument('source')
def importdata(source):
    wb = openpyxl.load_workbook(source)
    importmatches(wb["matches"])
    print("matches are imported success fully ")
    # wb=openpyxl.load_workbook('mockresult.xlsx')
    # importmocktest(wb.active)
    # print("marksdone are imported success fully ")

if __name__=='__main__':
    importdata()