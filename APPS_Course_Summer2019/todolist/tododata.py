import  click
import openpyxl
import os
import django
import todolist
import todoapp
from datetime import datetime

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'todolist.settings')
django.setup()


def importtodolist(worksheet):
    skip_flag = True
    for row in worksheet.iter_rows(values_only=True):
        print(row[0])
        if skip_flag:
            skip_flag = False
            continue
        try:
            Tdlist = todoapp.models.tdlist(name=row[0])
            Tdlist.save()
        except:
            print("Failed tuple:", row)

def importtodoitem(worksheet):
    skip_flag = True
    for row in worksheet.iter_rows(values_only=True):
        if skip_flag:
            skip_flag = False
            continue
        try:
            todo_list = todoapp.models.tdlist.objects.get(name=row[3])
            if (row[2] == None):
                Tditem = todoapp.models.tditem(description=row[0], due_date=row[1], completed=False,
                                               todolist=todo_list)
            elif (row[2] == 'TRUE'):
                Tditem = todoapp.models.tditem(description=row[0], due_date=row[1], completed=True, todolist=todo_list)
            else:
                Tditem = todoapp.models.tditem(description=row[0], due_date=row[1], completed=False, todolist=todo_list)
            Tditem.save()
        except:
            print("Failed tuple:", row)



@click.command()
@click.argument('source')
def importdata(source):
    wb = openpyxl.load_workbook(source)
    importtodolist(wb["list"])
    print("todolist are imported success fully ")
    importtodoitem(wb["item"])
    print("todoitems are imported success fully ")


if __name__=='__main__':
    importdata()